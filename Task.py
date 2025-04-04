import json
import requests
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

def scrape_website(url):
    """Scrapes the website and extracts UI elements."""
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    elements = {
        "buttons": [btn.get_text(strip=True) for btn in soup.find_all('button')],
        "links": [a['href'] for a in soup.find_all('a', href=True)],
        "input_fields": [inp['name'] if inp.has_attr('name') else 'Unnamed' for inp in soup.find_all('input')],
        "forms": [form['action'] if form.has_attr('action') else 'No action' for form in soup.find_all('form')]
    }
    
    with open("elements.json", "w") as f:
        json.dump(elements, f, indent=4)
    
    print("Scraping complete. Elements saved to elements.json")

def generate_test_cases():
    """Generates test cases based on elements.json."""
    with open("elements.json", "r") as f:
        elements = json.load(f)
    
    test_cases = [
        ["TC001", "Verify button click", "Click on a button", "Button should be clickable"],
        ["TC002", "Check form submission", "Fill input fields and submit form", "Form should be submitted successfully"],
        ["TC003", "Verify Sign Up Functionality", "Fill in the username and password fields with new credentials and Click the Sign Up button.", "Data is sent to /submit, and the user is redirected or shown a success message"]
    ]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Test Case ID", "Test Scenario", "Steps to Execute", "Expected Result"])
    for case in test_cases:
        ws.append(case)
    
    wb.save("test_cases.xlsx")
    print("Test cases saved to test_cases.xlsx")

def generate_selenium_scripts():
    """Generates Selenium scripts for test cases."""
    wb = openpyxl.load_workbook("test_cases.xlsx")
    ws = wb.active
    
    selenium_scripts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if "button" in row[1].lower():
            script = """
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://demoblaze.com")
driver.find_element(By.TAG_NAME, "button").click()
driver.quit()
            """.strip()
            selenium_scripts.append([row[0], script])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Test Case ID", "Python Selenium Code"])
    for script in selenium_scripts:
        ws.append(script)
    
    wb.save("test_scripts.xlsx")
    print("Selenium scripts saved to test_scripts.xlsx")

if __name__ == "__main__":
    scrape_website("https://demoblaze.com")
    generate_test_cases()
    generate_selenium_scripts()
